# utils/excel_exporter.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def export_treeview_to_excel(treeview, column_config, default_filename="export.xlsx"):
    """
    TkinterのTreeviewのデータを綺麗に装飾してExcelに出力する共通関数
    """
    import tkinter as tk
    from tkinter import filedialog, messagebox

    # 1. 出力するデータがあるかチェック
    items = treeview.get_children()
    if not items:
        messagebox.showwarning("出力不可", "出力するデータがありません。")
        return

    # 2. 保存先ダイアログ
    file_path = filedialog.asksaveasfilename(
        title="Excelファイルを出力",
        initialfile=default_filename,
        filetypes=[("Excelファイル", "*.xlsx")],
        defaultextension=".xlsx"
    )
    if not file_path:
        return

    try:
        # 3. ワークブック作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "データ一覧"

        # 4. ヘッダー（列名）の書き込み
        headers = [config["text"] for config in column_config.values()]
        ws.append(headers)

        # 5. 明細データの書き込み
        for item_id in items:
            row_values = treeview.item(item_id)["values"]
            ws.append(row_values)

        # --- 🎨 装飾処理 ---
        header_font = Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        data_font = Font(name="Meiryo UI", size=11)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = data_font
                # 詳細や備考の列（Treeviewの列設定に "details" や "remarks" が含まれる場合）は左揃え、他は中央
                # ここでは簡易的に、文字列が長い想定の列以外を中央揃えにします
                if cell.column_letter in (get_column_letter(7), get_column_letter(8)):
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment

        # 6. 列幅自動調整
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    cell_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    if cell_len > max_len:
                        max_len = cell_len
            ws.column_dimensions[col_letter].width = max(min(max_len + 3, 40), 10)

        # 7. 保存
        wb.save(file_path)
        messagebox.showinfo("完了", f"Excelファイルを正常に出力しました。\n\n保存先:\n{file_path}")

    except Exception as e:
        messagebox.showerror("出力エラー", f"Excelファイルの作成中にエラーが発生しました。\n\n原因: {e}")